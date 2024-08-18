import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def process_excel_file(input_file_path, output_file_path, columns_to_remove, job_label):
    # Load the original workbook
    original_workbook = openpyxl.load_workbook(input_file_path)

    # Create a new workbook as a copy of the original
    new_workbook = Workbook()

    # Copy sheet content from 'candidates' to the new workbook
    candidates_sheet = original_workbook['candidates']
    new_candidates_sheet = new_workbook.create_sheet('candidates')
    for row in candidates_sheet.iter_rows(values_only=True):
        new_candidates_sheet.append(row)

    # 1 Drop and rename
    new_candidates_sheet.delete_rows(1, 4)
    new_candidates_sheet['A1'] = 'first name'
    new_candidates_sheet['B1'] = 'last name'

    # 2 seperate first and last name
    for i, row in enumerate(new_candidates_sheet.iter_rows(min_row=2, max_row=new_candidates_sheet.max_row, values_only=True), 2):
        if row[0]:
            # Safe find
            if row[0].find(' ') > 0:
                first_name, last_name = row[0].split(' ', 1)
            else:
                first_name = row[0]
                last_name = ''

            new_candidates_sheet['A' + str(i)] = first_name
            new_candidates_sheet['B' + str(i)] = last_name

    # 3 drop columns from new sheet
    for i in range(0, 20):
        for col_index, col in enumerate(new_candidates_sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
            if (col[0] != None) and (col[0] in columns_to_remove):
                new_candidates_sheet.delete_cols(col_index)
                break

    # new_candidates_sheet.insert_cols(new_candidates_sheet.max_column)
    new_candidates_sheet.cell(
        row=1, column=new_candidates_sheet.max_column, value='notes')

    # Process 'candidates-updates' sheet
    updates_sheet = original_workbook['candidates-updates']

    # get list of update names
    names = []
    for col in updates_sheet.iter_cols(min_col=2, max_col=2, values_only=True):
        names = list(col)
        names = names[2:]

    for name in names:

        # Create notes string & build it
        notes = ""
        for row in updates_sheet.iter_rows(min_row=2, max_row=updates_sheet.max_row, values_only=True):
            r = list(row)
            if r[1] == name:
                notes += f"{r[4]} Created Note for on {r[5]}:\n{r[6]}\n\n"

        # append note to appropriate person in sheet
        for i, row in enumerate(new_candidates_sheet.iter_rows(min_row=2, max_row=new_candidates_sheet.max_row, values_only=True)):
            r = list(row)
            full_name = r[0] + " " + r[1]

            if (full_name == name):
                i2 = i+2
                notes2 = f"Old Notes from Monday for: {full_name}:\n\n" + notes
                new_candidates_sheet['I' + str(i2)] = notes2
                continue

    # Add job_title to each row
    new_candidates_sheet['J1'] = "Monday Job Group"
    for i, row in enumerate(new_candidates_sheet.iter_rows(min_row=2, max_row=new_candidates_sheet.max_row, values_only=True), start=2):
        new_candidates_sheet['J' + str(i)] = job_label

    # delete default worksheet
    d = new_workbook['Sheet']
    new_workbook.remove(d)

    # Save the changes to the new workbook
    new_workbook.save(output_file_path)


# Define columns to remove
columns_to_remove = ['Resume', 'Date Applied',
                     'Availability',  'Status 1', 'Status', 'Source',  'Location Map', 'Status', 'Cert. Status', 'Exp. Date', 'link to Clients', 'Recruiter', 'Message Status', 'Add To Favorites', 'Jobs', 'Item ID (auto generated)']  # Add your column names here

# Call the function with input and output file paths


job = "RADT/CADC/Tech"  # For Label in Job Diva
job_path = "RADT:CADC:TECH 2"  # For Local Path

path = f"/Users/victorrinaldi/Desktop/projects/auto_candidate/scripts/port_to_jobdiva/excel_files/{job_path}.xlsx"
outputpath = f'/Users/victorrinaldi/Desktop/projects/auto_candidate/scripts/port_to_jobdiva/output/{job_path}.xlsx'
process_excel_file(path,
                   outputpath, columns_to_remove, job_label=job)
